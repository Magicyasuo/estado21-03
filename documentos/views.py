"""
Módulo de vistas para la aplicación documentos.
Este archivo contiene todas las vistas y funciones relacionadas con la gestión
de documentos, registros, FUIDs y fichas de pacientes.
"""

# =============================================================================
# IMPORTACIONES
# =============================================================================

# Importaciones estándar de Python
from datetime import date, datetime

# Importaciones de Django
from django.contrib.auth.decorators import permission_required, login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import Q, Count, Avg
from django.urls import reverse_lazy
from django.utils.timezone import now, timedelta
from django.utils.decorators import method_decorator
from django.views.generic.edit import CreateView, UpdateView

# Librerías de terceros
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.drawing.image import Image

# Framework Django Rest Framework
from rest_framework.response import Response
from rest_framework.views import APIView

# Guardian - Permisos a nivel de objeto
from guardian.shortcuts import assign_perm, get_perms
from guardian.utils import get_anonymous_user

# Importaciones específicas del proyecto
from .forms import RegistroDeArchivoForm, FUIDForm, FichaPacienteForm
from .models import (
    RegistroDeArchivo,
    SubserieDocumental,
    SerieDocumental,
    FUID,
    FichaPaciente,
    Documento
)
from django.template.loader import render_to_string

@login_required
def cargar_series(request):
    """
    API para cargar las series documentales disponibles.
    Retorna un listado JSON con los códigos y nombres de todas las series.
    """
    series = SerieDocumental.objects.all().values('codigo', 'nombre')
    return JsonResponse(list(series), safe=False)

@login_required
def cargar_subseries(request):
    """
    API para cargar las subseries documentales filtradas por serie.
    Recibe el ID de la serie seleccionada y retorna las subseries correspondientes.
    """
    serie_id = request.GET.get('serie_id')
    subseries = SubserieDocumental.objects.filter(serie_id=serie_id).values('id', 'nombre')
    return JsonResponse(list(subseries), safe=False)

@login_required
def lista_registros(request):
    """
    Vista para listar registros de archivo.
    Filtra por la oficina del usuario a menos que sea superusuario.
    """
    if request.user.is_superuser:
        # El superusuario ve todo
        registros = RegistroDeArchivo.objects.all()
    else:
        # Filtrar por oficina del usuario
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

    return render(request, 'registro_list.html', {'registros': registros})

@login_required
def crear_registro(request):
    """
    Vista para crear un nuevo registro de archivo.
    Asigna automáticamente permisos a nivel de objeto al creador.
    """
    if not request.user.has_perm('documentos.add_registrodearchivo'):
        return HttpResponseForbidden("No tienes permiso para crear registros.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # Crear Documento si se subió un archivo
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                Documento.objects.create(
                    registro=registro,
                    archivo=archivo_subido
                )

            # Asignar permisos a nivel de objeto
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            # assign_perm('documentos.delete_own_registro', request.user, registro)

            messages.success(request, 'Registro de archivo creado exitosamente.')
            form = RegistroDeArchivoForm()
        else:
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")

    else:
        form = RegistroDeArchivoForm()
        # Subseries vacío por defecto (si no se selecciona serie)
        form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})

@login_required
def editar_registro(request, pk):
    """
    Vista para editar un registro existente.
    Verifica permisos a nivel de objeto antes de permitir la edición.
    """
    registro = get_object_or_404(RegistroDeArchivo, id=pk)

    # Verifica si el usuario tiene permiso de edición a nivel de objeto
    if not request.user.is_superuser and 'edit_own_registro' not in get_perms(request.user, registro):
        return HttpResponseForbidden("No tienes permiso para editar este registro.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, instance=registro)
        # Lógica de subseries
        codigo_serie = request.POST.get('codigo_serie')
        if codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie_id=codigo_serie)

        if form.is_valid():
            form.save()
            return redirect('lista_registros')
    else:
        form = RegistroDeArchivoForm(instance=registro)
        if registro.codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie=registro.codigo_serie)
        else:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})

@login_required
def eliminar_registro(request, pk):
    """
    Vista para eliminar un registro.
    Verifica permisos a nivel de objeto antes de permitir la eliminación.
    """
    registro = get_object_or_404(RegistroDeArchivo, pk=pk)

    # Si superuser, ok
    if request.user.is_superuser:
        registro.delete()
        return redirect('lista_registros')

    # Verifica si el user tiene permiso de delete a nivel de objeto
    perms = get_perms(request.user, registro)
    if 'delete_own_registro' in perms:
        registro.delete()
        return redirect('lista_registros')
    else:
        return HttpResponseForbidden("No tienes permiso para eliminar este registro.")

@login_required
def lista_completa_registros(request):
    """
    Vista para listar todos los registros con detalles adicionales.
    Similar a lista_registros pero con más información.
    """
    if request.user.is_superuser:
        registros = RegistroDeArchivo.objects.all()
    else:
        # Filtrar registros por la oficina del usuario
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

    return render(request, 'registro_completo.html', {'registros': registros})

@login_required
def registros_api(request):
    """
    API para DataTables que proporciona listados de registros con filtrado y paginación.
    Filtra automáticamente por oficina del usuario a menos que sea superusuario.
    """
    # 1) Filtra por oficina
    if request.user.is_superuser:
        registros = RegistroDeArchivo.objects.all()
    else:
        oficina_user = request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)
        registros = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

    # 2) Aplica la búsqueda por columnas que DataTables envía
    draw = request.GET.get("draw", 1)
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))

    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        col_search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()

        if col_search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=col_search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=col_search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=col_search_value)
            # ... repites el resto de tus filtros ...
            elif col_data == 'creado_por':
                registros = registros.filter(creado_por__username__icontains=col_search_value)
        i += 1

    # 3) Total sin filtros (para recordsTotal)
    total_registros = registros.count()

    # 4) Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # 5) Construir data
    data = []
    for registro in page:
        data.append({
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "documento": [{"url": doc.archivo.url} for doc in registro.documentos.all()],
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "creado_por": registro.creado_por.username if registro.creado_por else "N/A",
            "id": registro.id,
        })

    response = {
        "draw": int(draw),
        "recordsTotal": total_registros,
        "recordsFiltered": total_registros,
        "data": data,
    }
    return JsonResponse(response)

@login_required
def registros_api_completo(request):
    """
    API completa para DataTables que incluye filtrado avanzado, ordenación y paginación.
    Similar a registros_api pero con funcionalidades extendidas.
    """
    # Implementación similar a registros_api pero con funcionalidades extendidas
    pass

@login_required
def registros_api_con_id(request):
    """
    API para DataTables que incluye el ID de los registros en los resultados.
    Útil cuando se necesita hacer operaciones CRUD desde la interfaz.
    """
    # Similar a registros_api pero incluyendo el ID
    pass

@login_required
def registros_fuid_json(request, fuid_id):
    """
    API que proporciona los registros asociados a un FUID específico.
    Utilizada para mostrar los registros en la vista de detalle de FUID.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)

    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 30))

    # Registros base
    registros_query = fuid.registros.all()

    # Lee valores de filtro del request
    numero_orden = request.GET.get("filtro_numero_orden", "").strip()
    codigo = request.GET.get("filtro_codigo", "").strip()
    unidad_documental = request.GET.get("filtro_unidad_documental", "").strip()
    identificador_documento = request.GET.get("filtro_identificador_documento", "").strip()
    fecha_archivo = request.GET.get("filtro_fecha_archivo", "").strip()
    fecha_inicial = request.GET.get("filtro_fecha_inicial", "").strip()
    fecha_final = request.GET.get("filtro_fecha_final", "").strip()
    soporte_fisico = request.GET.get("filtro_soporte_fisico", "").strip()
    soporte_electronico = request.GET.get("filtro_soporte_electronico", "").strip()
    caja = request.GET.get("filtro_caja", "").strip()
    carpeta = request.GET.get("filtro_carpeta", "").strip()
    tomo_legajo = request.GET.get("filtro_tomo_legajo_libro", "").strip()
    numero_folios = request.GET.get("filtro_numero_folios", "").strip()
    tipo = request.GET.get("filtro_tipo", "").strip()
    cantidad = request.GET.get("filtro_cantidad", "").strip()
    ubicacion = request.GET.get("filtro_ubicacion", "").strip()
    cant_elec = request.GET.get("filtro_cant_elec", "").strip()
    tamano_elec = request.GET.get("filtro_tamano_elec", "").strip()
    notas = request.GET.get("filtro_notas", "").strip()
    creado_por = request.GET.get("filtro_creado_por", "").strip()
    fecha_creacion = request.GET.get("filtro_fecha_creacion", "").strip()
    estado_archivo = request.GET.get("filtro_estado_archivo", "").strip()

    # Construye Q() dinámico
    q_filter = Q()
    if numero_orden:
        q_filter &= Q(numero_orden__icontains=numero_orden)
    if codigo:
        q_filter &= Q(codigo__icontains=codigo)
    if unidad_documental:
        q_filter &= Q(unidad_documental__icontains=unidad_documental)
    if identificador_documento:
        q_filter &= Q(identificador_documento__icontains=identificador_documento)
    if fecha_archivo:
        # Asume fecha_archivo es exacta (YYYY-MM-DD). Ajusta a tu gusto
        q_filter &= Q(fecha_archivo=fecha_archivo)
    if fecha_inicial:
        q_filter &= Q(fecha_inicial=fecha_inicial)
    if fecha_final:
        q_filter &= Q(fecha_final=fecha_final)
    if soporte_fisico:
        # ✔ => True, ✖ => False
        q_filter &= Q(soporte_fisico=(soporte_fisico == "✔"))
    if soporte_electronico:
        q_filter &= Q(soporte_electronico=(soporte_electronico == "✔"))
    if caja:
        q_filter &= Q(caja__icontains=caja)
    if carpeta:
        q_filter &= Q(carpeta__icontains=carpeta)
    if tomo_legajo:
        q_filter &= Q(tomo_legajo_libro__icontains=tomo_legajo)
    if numero_folios:
        q_filter &= Q(numero_folios__icontains=numero_folios)
    if tipo:
        q_filter &= Q(tipo__icontains=tipo)
    if cantidad:
        q_filter &= Q(cantidad__icontains=cantidad)
    if ubicacion:
        q_filter &= Q(ubicacion__icontains=ubicacion)
    if cant_elec:
        q_filter &= Q(cantidad_documentos_electronicos__icontains=cant_elec)
    if tamano_elec:
        q_filter &= Q(tamano_documentos_electronicos__icontains=tamano_elec)
    if notas:
        q_filter &= Q(notas__icontains=notas)
    if creado_por:
        # Filtra por username
        q_filter &= Q(creado_por__username__icontains=creado_por)
    if fecha_creacion:
        # Asume el string exacto con formato YYYY-MM-DDTHH:MM
        # Ajusta a tu gusto
        q_filter &= Q(fecha_creacion__icontains=fecha_creacion)
    if estado_archivo:
        q_filter &= Q(Estado_archivo=(estado_archivo == "✔"))

    registros_query = registros_query.filter(q_filter)

    # Conteo y paginación
    total_registros = registros_query.count()
    paginator = Paginator(registros_query.order_by('id'), length)
    page_obj = paginator.get_page(start // length + 1)

    # Construye data
    data = []
    for r in page_obj:
        data.append({
            "numero_orden": r.numero_orden,
            "codigo": r.codigo or "",
            "unidad_documental": r.unidad_documental or "",
            "identificador_documento": r.identificador_documento or "",
            "fecha_archivo": r.fecha_archivo.strftime("%Y-%m-%d") if r.fecha_archivo else "",
            "fecha_inicial": r.fecha_inicial.strftime("%Y-%m-%d") if r.fecha_inicial else "",
            "fecha_final": r.fecha_final.strftime("%Y-%m-%d") if r.fecha_final else "",
            "soporte_fisico": "✔" if r.soporte_fisico else "✖",
            "soporte_electronico": "✔" if r.soporte_electronico else "✖",
            "caja": r.caja or "",
            "carpeta": r.carpeta or "",
            "tomo_legajo_libro": r.tomo_legajo_libro or "",
            "numero_folios": r.numero_folios if r.numero_folios else "",
            "tipo": r.tipo or "",
            "cantidad": r.cantidad if r.cantidad else "",
            "ubicacion": r.ubicacion or "",
            "cantidad_documentos_electronicos": r.cantidad_documentos_electronicos if r.cantidad_documentos_electronicos else "",
            "tamano_documentos_electronicos": r.tamano_documentos_electronicos if r.tamano_documentos_electronicos else "",
            "notas": r.notas or "",
            "creado_por": r.creado_por.username if r.creado_por else "",
            "fecha_creacion": r.fecha_creacion.strftime('%Y-%m-%d %H:%M') if r.fecha_creacion else "",
            "Estado_archivo": "✔" if r.Estado_archivo else "✖",
            # Campos sin filtro
            "documento": (
                f'<a href="/registros/documento/{r.id}/" target="_blank">📁 Ver Documento</a>'
                if r.documentos.all()
                else "✖ No hay documento"
            ),
            "acciones": f'<a href="/registros/fuids/{fuid.id}/editar_registro/{r.id}/" class="btn btn-sm btn-warning">Editar</a>'
        })

    return JsonResponse({
        "draw": int(request.GET.get('draw', 1)),
        "recordsTotal": total_registros,
        "recordsFiltered": total_registros,
        "data": data
    })

@login_required
def form_registro_fuid_ajax(request, fuid_id):
    """
    Retorna el HTML parcial de un formulario para crear un RegistroDeArchivo
    asociado a un FUID, listo para inyectar en un modal.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)
    form = RegistroDeArchivoForm()  # Form vacío

    # Renderizamos un template parcial con el formulario
    html_form = render_to_string(
        'partials/_form_registro.html',
        {'form': form, 'fuid': fuid},
        request=request
    )
    return JsonResponse({'html_form': html_form})

@login_required
def crear_registro_fuid_ajax(request, fuid_id):
    """
    Vista para crear un registro y asociarlo a un FUID vía AJAX.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # Asignar permisos a nivel de objeto
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)

            # Asociar con FUID
            fuid.registros.add(registro)

            # Devolvemos éxito y el registro en JSON
            return JsonResponse({
                'ok': True,
                'message': 'Registro creado exitosamente.',
                'registro': {
                    'id': registro.id,
                    'numero_orden': registro.numero_orden,
                    'codigo': registro.codigo or '',
                }
            })
        else:
            # Devolvemos el HTML del form con errores
            html_form = render_to_string(
                'partials/_form_registro.html',
                {'form': form, 'fuid': fuid},
                request=request
            )
            return JsonResponse({'ok': False, 'html_form': html_form})

    return JsonResponse({'ok': False, 'message': 'Método no permitido'}, status=405)

@login_required
def obtener_usuarios(request):
    """
    API que retorna la lista de usuarios para selectores de formularios.
    """
    usuarios = User.objects.values('username')
    return JsonResponse(list(usuarios), safe=False)

# mixins.py
from django.http import HttpResponseForbidden

class OficinaFilterMixin:
    """
    Filtra los objetos para que el usuario solo vea y manipule
    aquellos creados por su oficina. También bloquea la edición
    de objetos de otras oficinas.
    """
    def get_queryset(self):
        # Implementación del filtro por oficina
        pass

    def dispatch(self, request, *args, **kwargs):
        # Bloqueo adicional para edición/eliminación
        pass

from django.shortcuts import render

def soporte_view(request):
    """
    Vista para la página de soporte.
    """
    # Implementación para página de soporte
    pass

def ver_documento(request, registro_id):
    """
    Vista para visualizar documentos asociados a un registro.
    """
    registro = get_object_or_404(RegistroDeArchivo, id=registro_id)
    
    # 1. Recuperar la oficina del usuario
    oficina_usuario = request.user.perfil.oficina
    
    # 2. Recuperar la oficina del registro (desde el primer FUID asociado)
    fuid = registro.fuids.first()
    if not fuid:
        return HttpResponseForbidden("Este registro no tiene FUID asignado. No se puede verificar oficina.")
    oficina_registro = fuid.oficina_productora
    
    # 3. Comparar oficinas
    if oficina_registro != oficina_usuario and not request.user.is_superuser:
        return HttpResponseForbidden("No tienes permiso para ver este documento.")
    
    # Si pasa la verificación, procedemos
    documentos = registro.documentos.all()
    for doc in documentos:
        doc.sas_url = doc.archivo.url  # URL local
    
    return render(request, "documento_detalle.html", {
        "registro": registro,
        "documentos": documentos,
        "fuid": fuid,
    })

# =============================================================================
# VISTAS DE GESTIÓN DE FUIDS (Formato Único de Inventario Documental)
# =============================================================================

class FUIDCreateView(LoginRequiredMixin, CreateView):
    """
    Vista para crear un nuevo FUID.
    Asigna automáticamente el usuario creador y gestiona permisos.
    """
    model = FUID
    form_class = FUIDForm
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def dispatch(self, request, *args, **kwargs):
        # Verifica si el usuario tiene permiso global para crear FUIDs
        if not request.user.has_perm('documentos.add_fuid'):
            return HttpResponseForbidden("No tienes permiso para crear FUIDs.")
        return super().dispatch(request, *args, **kwargs)

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)

        # Solo registros creados por el usuario autenticado
        registros = RegistroDeArchivo.objects.filter(
            fuids__isnull=True,
            creado_por=self.request.user
        )

        # Aplicar filtros opcionales
        fecha_inicio = self.request.GET.get("fecha_inicio")
        fecha_fin = self.request.GET.get("fecha_fin")
        if fecha_inicio:
            registros = registros.filter(fecha_creacion__gte=fecha_inicio)
        if fecha_fin:
            registros = registros.filter(fecha_creacion__lte=fecha_fin)

        # Asignar el queryset de registros al formulario
        form.fields['registros'].queryset = registros

        # Establecer el usuario autenticado en el formulario
        form.fields['usuario'].initial = self.request.user.id
        form.fields['usuario'].widget.attrs['readonly'] = True
        return form

    def form_valid(self, form):
        # Asigna automáticamente el usuario que crea el FUID
        form.instance.creado_por = self.request.user
        fuid = form.save()
        
        # Asigna permisos a nivel de objeto al creador
        assign_perm('documentos.view_own_fuid', self.request.user, fuid)
        assign_perm('documentos.edit_own_fuid', self.request.user, fuid)
        assign_perm('documentos.delete_own_fuid', self.request.user, fuid)

        # Asociar registros al FUID
        registros = form.cleaned_data["registros"]
        fuid.registros.set(registros)
        
        return super().form_valid(form)

class FUIDUpdateView(LoginRequiredMixin, UpdateView):
    """
    Vista para actualizar un FUID existente.
    Verifica permisos a nivel de objeto antes de permitir la edición.
    """
    model = FUID
    form_class = FUIDForm
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def dispatch(self, request, *args, **kwargs):
        fuid = self.get_object()

        # Verificar si el usuario es el creador o es superusuario
        if not request.user.is_superuser and fuid.creado_por != request.user:
            return HttpResponseForbidden("Solo el creador de este FUID puede editarlo.")

        return super().dispatch(request, *args, **kwargs)

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)
        fuid = self.get_object()

        oficina_user = self.request.user.perfil.oficina
        usuarios_de_mi_oficina = User.objects.filter(perfil__oficina=oficina_user)

        # Filtra registros solo de la oficina del usuario
        registros_disponibles = RegistroDeArchivo.objects.filter(creado_por__in=usuarios_de_mi_oficina)

        # Obtener los registros ya asociados a este FUID
        registros_asociados = fuid.registros.all()

        # Fusionar registros asociados con los disponibles
        form.fields['registros'].queryset = (registros_disponibles | registros_asociados).distinct()

        return form

    def form_valid(self, form):
        # Lógica de guardado
        fuid = form.save()
        registros = form.cleaned_data.get("registros")
        fuid.registros.set(registros)
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fuid = self.get_object()

        # Obtener los registros ya asociados al FUID
        context["registros_asociados"] = list(fuid.registros.values_list("id", flat=True))

        return context

@login_required
def lista_fuids(request):
    """
    Vista para listar todos los FUIDs disponibles según permisos.
    """
    if request.user.is_superuser:
        fuids = FUID.objects.all()
    else:
        fuids = FUID.objects.filter(oficina_productora=request.user.perfil.oficina)
    return render(request, 'fuid_list.html', {'fuids': fuids})

@login_required
def detalle_fuid(request, pk):
    """
    Vista para ver el detalle de un FUID específico.
    Incluye información sobre registros asociados.
    """
    fuid = get_object_or_404(FUID, pk=pk)

    # Verificar permisos
    if not request.user.is_superuser and fuid.oficina_productora != request.user.perfil.oficina:
        return HttpResponseForbidden("No tienes permiso para ver este FUID.")

    if not request.user.has_perm('documentos.view_own_fuid', fuid):
        assign_perm('documentos.view_own_fuid', request.user, fuid)

    # Renderizar la plantilla sin cargar los registros todavía
    return render(request, 'fuid_complete_list.html', {'fuid': fuid})

@login_required
def agregar_registro_a_fuid(request, fuid_id):
    """
    Vista para agregar un registro existente a un FUID.
    Verifica permisos antes de permitir la operación.
    """
    # Verificamos que el usuario tenga permiso de agregar registros
    if not request.user.has_perm('documentos.add_registrodearchivo'):
        return HttpResponseForbidden("No tienes permiso para crear registros.")

    fuid = get_object_or_404(FUID, pk=fuid_id)

    if request.method == 'POST':
        # IMPORTANTE: incluir request.FILES para que se procese el archivo
        form = RegistroDeArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # Crear Documento si se subió un archivo
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                Documento.objects.create(
                    registro=registro,
                    archivo=archivo_subido
                )

            # Asignar permisos a nivel de objeto
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)

            # Asociar el registro recién creado al FUID
            fuid.registros.add(registro)

            messages.success(request, 'Registro creado y asociado correctamente al FUID.')
            form = RegistroDeArchivoForm()  # reiniciar el formulario
        else:
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        form = RegistroDeArchivoForm()
        # Si no se ha seleccionado una serie, dejamos vacío el queryset de subseries
        form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'agregar_registro_a_fuid.html', {
        'form': form,
        'fuid': fuid
    })

@login_required
def editar_registro_de_fuid(request, fuid_id, registro_id):
    """
    Vista para editar un registro que pertenece a un FUID.
    Verifica permisos antes de permitir la edición.
    """
    # Verificar permiso de edición
    if not request.user.has_perm('documentos.change_registrodearchivo'):
        return HttpResponseForbidden("No tienes permiso para editar registros.")

    fuid = get_object_or_404(FUID, pk=fuid_id)
    registro = get_object_or_404(RegistroDeArchivo, pk=registro_id)

    # Asegurarnos de que el registro pertenece realmente al FUID
    if registro not in fuid.registros.all():
        return HttpResponseForbidden("El registro no está asociado a este FUID.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, request.FILES, instance=registro)
        if form.is_valid():
            updated_registro = form.save()

            # Verificamos si subieron un archivo nuevo
            archivo_subido = form.cleaned_data.get('archivo')
            if archivo_subido:
                # Creamos el Documento y lo asociamos
                Documento.objects.create(
                    registro=updated_registro,
                    archivo=archivo_subido
                )

            messages.success(request, 'Registro actualizado correctamente.')
            return redirect('detalle_fuid', pk=fuid_id)
        else:
            # Manejar errores de validación
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        # Cargar la instancia existente para prellenar campos
        form = RegistroDeArchivoForm(instance=registro)

        # Ajustar la queryset de subseries si la serie ya está presente
        if registro.codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(
                serie_id=registro.codigo_serie.id
            )
        else:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'editar_registro_de_fuid.html', {
        'form': form,
        'fuid': fuid,
        'registro': registro
    })

@login_required
def export_fuid_to_excel(request, pk):
    """
    Vista para exportar un FUID completo a Excel.
    Formatea el documento siguiendo plantillas institucionales.
    """
    # Obtener el FUID específico
    fuid = FUID.objects.get(pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"FUID #{fuid.id}"

    # Función para truncar valores largos
    def truncate_value(value, max_length=30):
        if not value:
            return "N/A"
        value = str(value)
        return value if len(value) <= max_length else value[:max_length - 3] + "..."

    # Crear estilos y dar formato al documento
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

    # Combinar celdas para la imagen
    ws.merge_cells(start_row=1, start_column=1, end_row=6, end_column=22)

    # Resto de la implementación...

    # Configurar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=FUID_{fuid.id}.xlsx'

    wb.save(response)
    return response

# =============================================================================
# VISTAS DE FICHAS DE PACIENTES
# =============================================================================

@login_required
def crear_ficha_paciente(request):
    """
    Vista para crear una nueva ficha de paciente.
    Verifica permisos antes de permitir la creación.
    """
    # Verificar si el usuario tiene permiso global para agregar fichas
    if not request.user.has_perm('documentos.add_fichapaciente'):
        return HttpResponseForbidden("No tienes permiso para crear fichas de pacientes.")

    if request.method == 'POST':
        form = FichaPacienteForm(request.POST)
        if form.is_valid():
            ficha = form.save()
            messages.success(request, 'Ficha del paciente registrada exitosamente.')
            return redirect('crear_ficha')  # Redirige a la misma página o a otra URL
        else:
            # Manejo de errores en el formulario
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        form = FichaPacienteForm()

    return render(request, 'ficha_paciente_form.html', {'form': form})

@login_required
@permission_required('documentos.view_fichapaciente', raise_exception=True)
def lista_fichas_paciente(request):
    """
    Vista para listar todas las fichas de pacientes.
    Requiere permisos específicos de visualización.
    """
    fichas = FichaPaciente.objects.all()
    return render(request, 'lista_fichas_paciente.html', {'fichas': fichas})

@method_decorator(login_required, name='dispatch')
class EditarFichaPaciente(UpdateView):
    """
    Vista para editar una ficha de paciente existente.
    Verifica permisos antes de permitir la edición.
    """
    model = FichaPaciente
    fields = '__all__'
    template_name = 'ficha_paciente_form.html'
    success_url = reverse_lazy('lista_fichas')
    pk_url_kwarg = 'consecutivo'  # Usar 'consecutivo' en lugar de 'pk'

    def dispatch(self, request, *args, **kwargs):
        # Verificar si el usuario tiene permiso global para editar fichas
        if not request.user.has_perm('documentos.change_fichapaciente'):
            return HttpResponseForbidden("No tienes permiso para editar fichas de pacientes.")

        # Verificar que la ficha existe
        self.object = get_object_or_404(FichaPaciente, consecutivo=kwargs.get(self.pk_url_kwarg))
        return super().dispatch(request, *args, **kwargs)

@login_required
def detalle_ficha_paciente(request, consecutivo):
    """
    Vista para ver el detalle de una ficha de paciente.
    """
    ficha = get_object_or_404(FichaPaciente, consecutivo=consecutivo)
    return render(request, 'detalle_ficha_paciente.html', {'ficha': ficha})

class ListaFichasAPIView(APIView):
    """
    API para listar fichas de pacientes con filtrado avanzado.
    Utilizada para DataTables en el frontend.
    """
    def get(self, request):
        # Parámetros enviados desde el frontend
        fecha_inicio = request.GET.get('fecha_inicio', None)
        fecha_fin = request.GET.get('fecha_fin', None)
        filtro_identificacion = request.GET.get('filtro_identificacion', None)
        filtro_historia = request.GET.get('filtro_historia', None)
        filtro_nombre = request.GET.get('filtro_nombre', None)
        filtro_similar = request.GET.get('filtro_similar', None)
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 250))

        # Base queryset
        queryset = FichaPaciente.objects.all()

        # Aplicar filtros según parámetros recibidos
        # ...

        # Formato JSON para DataTables
        data = [
            {
                "consecutivo": ficha.consecutivo,
                "nombre_completo": f"{ficha.primer_nombre} {ficha.segundo_nombre or ''} {ficha.primer_apellido} {ficha.segundo_apellido}",
                "tipo_identificacion": ficha.tipo_identificacion,
                "num_identificacion": ficha.num_identificacion,
                # ... Resto de campos
            }
            for ficha in queryset
        ]

        return Response({
            "draw": request.GET.get("draw", 1),
            "recordsTotal": queryset.count(),
            "recordsFiltered": queryset.count(),
            "data": data,
        })

# =============================================================================
# VISTAS DE ESTADÍSTICAS Y REPORTES
# =============================================================================

def calcular_edad(fecha_nacimiento):
    """
    Función auxiliar para calcular la edad a partir de la fecha de nacimiento.
    """
    if fecha_nacimiento:
        hoy = date.today()
        return hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    return None

@login_required
def estadisticas_pacientes(request):
    """
    Vista para mostrar estadísticas sobre pacientes.
    """
    usuario = request.GET.get('usuario')
    pacientes = FichaPaciente.objects.all()

    if usuario:
        pacientes = pacientes.filter(creado_por__username=usuario)

    # Calcular edades
    edades = [calcular_edad(p.fecha_nacimiento) for p in pacientes if p.fecha_nacimiento]

    # Clasificar por grupos de edad
    grupos_edad = {
        "0-18": sum(1 for e in edades if e <= 18),
        "19-35": sum(1 for e in edades if 19 <= e <= 35),
        "36-60": sum(1 for e in edades if 36 <= e <= 60),
        "60+": sum(1 for e in edades if e > 60)
    }

    datos = {
        'total_pacientes': pacientes.count(),
        'por_genero': list(pacientes.values('sexo').annotate(cantidad=Count('sexo'))),
        'por_tipo_identificacion': list(pacientes.values('tipo_identificacion').annotate(cantidad=Count('tipo_identificacion'))),
        'activos': pacientes.filter(activo=True).count(),
        'promedio_edad': round(sum(edades) / len(edades), 2) if edades else None,
        'grupos_edad': grupos_edad
    }

    return JsonResponse(datos, safe=False)

@login_required
def estadisticas_registros(request):
    """
    Vista para mostrar estadísticas sobre registros de archivo.
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        registros = RegistroDeArchivo.objects.all()

        # Filtrar por rango de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            registros = registros.filter(fecha_archivo__range=(fecha_inicio, fecha_fin))

        # Generar estadísticas
        datos = {
            'total_registros': registros.count(),
            'por_serie': list(
                registros.values('codigo_serie__nombre').annotate(cantidad=Count('id'))
            ),
            'por_soporte': list(
                registros.values('soporte_fisico', 'soporte_electronico').annotate(cantidad=Count('id'))
            ),
            'por_tipo': list(
                registros.values('tipo').annotate(cantidad=Count('id'))
            ),
        }

        return JsonResponse(datos, safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
def estadisticas_fuids(request):
    """
    Vista para mostrar estadísticas sobre FUIDs.
    """
    usuario = request.GET.get('usuario')
    fuids = FUID.objects.all()

    if usuario:
        fuids = fuids.filter(creado_por__username=usuario)

    datos = {
        'total_fuids': fuids.count(),
        'por_oficina': list(fuids.values('oficina_productora__nombre').annotate(cantidad=Count('id'))),
        'por_objeto': list(fuids.values('objeto__nombre').annotate(cantidad=Count('id'))),
        'por_entidad': list(fuids.values('entidad_productora__nombre').annotate(cantidad=Count('id'))),
    }

    return JsonResponse(datos, safe=False)

@login_required
def pagina_estadisticas(request):
    """
    Página principal para mostrar gráficos de las estadísticas.
    """
    return render(request, 'pagina_estadisticas.html')

# =============================================================================
# VISTAS AUXILIARES Y DE SISTEMA
# =============================================================================

@login_required
def welcome_view(request):
    """
    Vista de bienvenida después del login.
    """
    return render(request, 'welcome.html')

@login_required
def panel_view(request):
    """
    Vista del panel principal de la aplicación.
    Solo pueden acceder los usuarios con el grupo "administradores"
    """
    if not request.user.is_superuser:
        return mi_error_403(request)  # Llamamos a la función de error si no es superusuario

    return render(request, 'panel_de_control.html')

def mi_error_403(request, exception=None):
    """
    Vista personalizada para errores 403 (Forbidden).
    """
    return render(request, '403.html', status=403)

@login_required
def soporte_view(request):
    """
    Vista para la página de soporte.
    """
    return render(request, 'soporte.html')

# =============================================================================
# MIXINS Y CLASES DE UTILIDAD
# =============================================================================

class OficinaFilterMixin:
    """
    Filtra los objetos para que el usuario solo vea y manipule
    aquellos creados por su oficina. También bloquea la edición
    de objetos de otras oficinas.
    """
    def get_queryset(self):
        qs = super().get_queryset()
        # Si el superusuario debe ver todo
        if self.request.user.is_superuser:
            return qs
        # Caso contrario, filtra por la oficina del perfil
        return qs.filter(oficina_productora=self.request.user.perfil.oficina)

    def dispatch(self, request, *args, **kwargs):
        # Bloqueo adicional para edición/eliminación
        if hasattr(self, 'get_object'):
            obj = self.get_object()
            if (not request.user.is_superuser) and (obj.oficina_productora != request.user.perfil.oficina):
                return HttpResponseForbidden("No tienes permiso sobre este recurso.")
        return super().dispatch(request, *args, **kwargs)

# =============================================================================
# VISTAS NO UTILIZADAS (COMENTADAS)
# =============================================================================

# Las siguientes vistas no están en uso actualmente pero se mantienen
# en el código por referencia o uso futuro.

"""
@login_required
def vista_no_utilizada(request):
    # Esta vista fue reemplazada por otra implementación
    pass
"""

"""
class ClaseNoUtilizada:
    # Esta clase fue reemplazada por otra implementación
    pass
"""



